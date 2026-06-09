"""Storage, indexing and retrieval for construction daily reports.

Each report lives in its own folder under ``reports/`` with a structured
``report.json``, generated Word/PDF exports, and a ``source_files/`` directory
holding the raw notes/emails/PDFs the report was built from. A master
``index.json`` tracks every report for fast listing.

    reports/
      index.json
      DR-001_2026-06-07/
        report.json
        DR-001_2026-06-07.docx
        DR-001_2026-06-07.pdf
        source_files/
          morning_notes.txt
"""

from __future__ import annotations

import json
import os
import shutil
from datetime import datetime

try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass

REPORTS_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "reports"
)
INDEX_PATH = os.path.join(REPORTS_DIR, "index.json")

# Report fields the LLM can populate (used for defaults and validation).
REPORT_TEXT_FIELDS = (
    "manpower",
    "equipment",
    "work_performed",
    "delays_issues",
    "safety_observations",
    "visitors",
)


def _now() -> str:
    """Return the current local time as an ISO string (seconds precision)."""
    return datetime.now().isoformat(timespec="seconds")


def _ensure_dirs() -> None:
    """Make sure the reports directory exists."""
    os.makedirs(REPORTS_DIR, exist_ok=True)


def _load_index() -> dict:
    """Load index.json, returning a fresh structure if missing/corrupt."""
    if not os.path.isfile(INDEX_PATH):
        return {"last_report_number": 0, "reports": []}
    try:
        with open(INDEX_PATH, "r", encoding="utf-8") as handle:
            data = json.load(handle)
    except (json.JSONDecodeError, OSError) as error:
        print(f"[reports] Could not read index.json: {error}")
        return {"last_report_number": 0, "reports": []}
    data.setdefault("last_report_number", 0)
    data.setdefault("reports", [])
    return data


def _save_index(index: dict) -> None:
    """Persist index.json."""
    _ensure_dirs()
    try:
        with open(INDEX_PATH, "w", encoding="utf-8") as handle:
            json.dump(index, handle, indent=2, ensure_ascii=False)
    except OSError as error:
        print(f"[reports] Failed to write index.json: {error}")


def get_report_folder(report_id: str) -> str:
    """Return the absolute folder path for ``report_id``."""
    return os.path.join(REPORTS_DIR, report_id)


def _report_json_path(report_id: str) -> str:
    """Return the path to a report's report.json."""
    return os.path.join(get_report_folder(report_id), "report.json")


def get_next_report_id(date_str: str) -> tuple[str, str]:
    """Return the next (report_number, id) pair without persisting it.

    ``report_number`` is ``DR-NNN`` and ``id`` is ``DR-NNN_<date>``.
    """
    index = _load_index()
    number = index.get("last_report_number", 0) + 1
    report_number = f"DR-{number:03d}"
    return report_number, f"{report_number}_{date_str}"


def create_report(date_str: str, settings: dict) -> dict:
    """Create a new report folder + report.json and register it in the index."""
    _ensure_dirs()
    index = _load_index()
    number = index.get("last_report_number", 0) + 1
    report_number = f"DR-{number:03d}"
    report_id = f"{report_number}_{date_str}"

    folder = get_report_folder(report_id)
    os.makedirs(os.path.join(folder, "source_files"), exist_ok=True)

    report = {
        "id": report_id,
        "report_number": report_number,
        "date": date_str,
        "project_name": settings.get("project_name", ""),
        "prepared_by": settings.get("prepared_by", ""),
        "weather_morning": "",
        "weather_afternoon": "",
        "weather_auto_fetched": False,
        "manpower": "",
        "equipment": "",
        "work_performed": "",
        "delays_issues": "",
        "safety_observations": "",
        "visitors": "",
        "source_files": [],
        "llm_extraction_used": False,
        "created_at": _now(),
        "last_edited_at": _now(),
        "status": "draft",
    }
    _write_report_json(report)

    index["last_report_number"] = number
    index["reports"].append(_index_entry(report))
    _save_index(index)
    return report


def _write_report_json(report: dict) -> None:
    """Write report.json into the report's folder."""
    path = _report_json_path(report["id"])
    os.makedirs(os.path.dirname(path), exist_ok=True)
    try:
        with open(path, "w", encoding="utf-8") as handle:
            json.dump(report, handle, indent=2, ensure_ascii=False)
    except OSError as error:
        print(f"[reports] Failed to write report.json: {error}")


def _index_entry(report: dict) -> dict:
    """Build the lightweight index entry for a report dict."""
    folder = get_report_folder(report["id"])
    docx = os.path.join(folder, f"{report['id']}.docx")
    pdf = os.path.join(folder, f"{report['id']}.pdf")
    return {
        "id": report["id"],
        "report_number": report["report_number"],
        "date": report["date"],
        "folder": folder,
        "created_at": report.get("created_at", ""),
        "last_edited_at": report.get("last_edited_at", ""),
        "has_docx": os.path.isfile(docx),
        "has_pdf": os.path.isfile(pdf),
        "source_file_count": len(report.get("source_files", [])),
        "weather_auto": report.get("weather_auto_fetched", False),
        "status": report.get("status", "draft"),
    }


def load_report(report_id: str) -> dict:
    """Load and return a report's report.json."""
    path = _report_json_path(report_id)
    if not os.path.isfile(path):
        raise FileNotFoundError(f"No report found: {report_id}")
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def save_report(report: dict) -> None:
    """Persist an edited report and refresh its index entry."""
    report["last_edited_at"] = _now()
    _write_report_json(report)

    index = _load_index()
    entry = _index_entry(report)
    found = False
    for i, existing in enumerate(index["reports"]):
        if existing["id"] == report["id"]:
            index["reports"][i] = entry
            found = True
            break
    if not found:
        index["reports"].append(entry)
    _save_index(index)


def list_reports() -> list[dict]:
    """Return all index entries sorted by date (newest first)."""
    index = _load_index()
    return sorted(
        index.get("reports", []),
        key=lambda r: (r.get("date", ""), r.get("id", "")),
        reverse=True,
    )


def refresh_index_entry(report_id: str) -> None:
    """Recompute and store an index entry (e.g. after exporting files)."""
    report = load_report(report_id)
    save_report(report)


# --------------------------------------------------------------------------- #
# Source files
# --------------------------------------------------------------------------- #
def _extract_text(path: str) -> str:
    """Extract plain text from a source file by extension; '' if unsupported."""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".txt":
            try:
                with open(path, "r", encoding="utf-8") as handle:
                    return handle.read()
            except UnicodeDecodeError:
                with open(path, "r", encoding="latin-1") as handle:
                    return handle.read()
        if ext == ".pdf":
            from pypdf import PdfReader

            reader = PdfReader(path)
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        if ext == ".docx":
            from docx import Document

            document = Document(path)
            return "\n".join(p.text for p in document.paragraphs)
        if ext == ".eml":
            import email
            from email import policy

            with open(path, "rb") as handle:
                message = email.message_from_binary_file(handle, policy=policy.default)
            subject = message.get("Subject", "")
            body = ""
            if message.is_multipart():
                for part in message.walk():
                    if part.get_content_type() == "text/plain":
                        body = part.get_content()
                        break
            else:
                body = message.get_content()
            return f"Subject: {subject}\n\n{body}"
        if ext in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            sheets: list[str] = []
            for sheet in workbook.worksheets:
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None and str(c).strip()]
                    if cells:
                        rows.append(" | ".join(cells))
                if rows:
                    sheets.append(f"[Sheet: {sheet.title}]\n" + "\n".join(rows))
            workbook.close()
            return "\n\n".join(sheets)
        if ext == ".csv":
            import csv

            try:
                handle = open(path, "r", encoding="utf-8", newline="")
            except UnicodeDecodeError:
                handle = open(path, "r", encoding="latin-1", newline="")
            with handle:
                rows = [
                    " | ".join(str(c) for c in row if str(c).strip())
                    for row in csv.reader(handle)
                ]
            return "\n".join(r for r in rows if r)
    except Exception as error:  # noqa: BLE001 - never let one bad file crash.
        print(f"[reports] Could not extract {os.path.basename(path)}: {error}")
        return ""
    return ""


def upload_source_file(report_id: str, filepath: str) -> dict:
    """Copy ``filepath`` into the report's source_files folder and index it.

    Returns metadata: ``{filename, type, size_bytes, extracted_chars}``.
    """
    folder = os.path.join(get_report_folder(report_id), "source_files")
    os.makedirs(folder, exist_ok=True)
    filename = os.path.basename(filepath)
    dest = os.path.join(folder, filename)
    shutil.copyfile(filepath, dest)

    ext = os.path.splitext(filename)[1].lower().lstrip(".") or "unknown"
    extracted = _extract_text(dest)
    size = os.path.getsize(dest)

    report = load_report(report_id)
    if filename not in report["source_files"]:
        report["source_files"].append(filename)
    save_report(report)

    return {
        "filename": filename,
        "type": ext,
        "size_bytes": size,
        "extracted_chars": len(extracted),
    }


def remove_source_file(report_id: str, filename: str) -> None:
    """Delete a single source file and drop it from report.json."""
    base = os.path.basename(filename)
    path = os.path.join(get_report_folder(report_id), "source_files", base)
    if os.path.isfile(path):
        os.remove(path)
    report = load_report(report_id)
    report["source_files"] = [f for f in report["source_files"] if f != base]
    save_report(report)


def get_source_texts(report_id: str) -> str:
    """Combine all source files into one headed string for the LLM."""
    folder = os.path.join(get_report_folder(report_id), "source_files")
    if not os.path.isdir(folder):
        return ""
    chunks: list[str] = []
    for name in sorted(os.listdir(folder)):
        path = os.path.join(folder, name)
        if os.path.isfile(path):
            text = _extract_text(path)
            chunks.append(f"=== {name} ===\n{text}")
    return "\n\n".join(chunks)


def mark_finalized(report_id: str) -> None:
    """Set a report's status to 'finalized' in report.json and the index."""
    report = load_report(report_id)
    report["status"] = "finalized"
    save_report(report)


def delete_report(report_id: str) -> None:
    """Remove a report's folder and its index entry."""
    folder = get_report_folder(report_id)
    if os.path.isdir(folder):
        shutil.rmtree(folder, ignore_errors=True)
    index = _load_index()
    index["reports"] = [r for r in index["reports"] if r["id"] != report_id]
    _save_index(index)


def find_today_draft(date_str: str) -> dict | None:
    """Return today's draft report dict if one exists, else None."""
    for entry in list_reports():
        if entry["date"] == date_str and entry["status"] == "draft":
            try:
                return load_report(entry["id"])
            except FileNotFoundError:
                continue
    return None


if __name__ == "__main__":
    print("--- daily_report_store self-test ---")
    settings = {"project_name": "Test Project", "prepared_by": "Test PM"}
    rep = create_report(datetime.now().strftime("%Y-%m-%d"), settings)
    print("Created:", rep["id"])

    rep["work_performed"] = "Poured footings at grid A-C."
    save_report(rep)
    print("Saved edit; last_edited_at:", load_report(rep["id"])["last_edited_at"])

    print("List count:", len(list_reports()))
    mark_finalized(rep["id"])
    print("Status:", load_report(rep["id"])["status"])

    delete_report(rep["id"])
    print("Deleted; list count:", len(list_reports()))
