"""Industrial-Slate daily-report Excel workbook generator.

Builds a three-sheet construction daily report (Labor & Progress, Machinery &
Equipment, HSE & Security) styled with the "Industrial Slate" theme: charcoal
banners, steel-slate section bars, muted-slate table headers, zebra striping,
soft status highlights, live formulas, number formatting and auto-fit columns.

Run standalone to emit a sample workbook, or call
:func:`build_daily_report_excel` to render one for a specific report.
"""

from __future__ import annotations

import os
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class Palette:
    """Industrial Slate colour palette (hex, no leading '#')."""

    CHARCOAL = "2D3748"
    STEEL = "4A5568"
    MUTED = "718096"
    ZEBRA = "F8FAFC"
    SUMMARY = "EDF2F7"
    WHITE = "FFFFFF"

    # Soft status tones with matching text colours.
    MINT = "C6F6D5"
    MINT_TEXT = "22543D"
    CRIMSON = "FED7D7"
    CRIMSON_TEXT = "742A2A"
    AMBER = "FEEBC8"
    AMBER_TEXT = "744210"
    SLATE_LIGHT = "E2E8F0"
    SLATE_LIGHT_TEXT = "4A5568"
    BLUE = "BEE3F8"
    BLUE_TEXT = "2A4365"
    TEAL = "B2F5EA"
    TEAL_TEXT = "234E52"


class DailyReportExcelBuilder:
    """Render the styled daily-report workbook using openpyxl."""

    FONT_NAME = "Segoe UI"

    def __init__(self, report: Optional[dict] = None) -> None:
        """Store report metadata and prepare an empty workbook."""
        self.report = report or {}
        self.workbook = Workbook()
        self._thin = Side(style="thin", color="CBD5E0")
        self._double = Side(style="double", color=Palette.CHARCOAL)
        self._border = Border(
            left=self._thin, right=self._thin, top=self._thin, bottom=self._thin
        )

    # --- low-level styling helpers --------------------------------------
    def _font(
        self, size: int = 10, bold: bool = False, color: str = "000000"
    ) -> Font:
        """Return a Segoe UI font with the given size/weight/colour."""
        return Font(name=self.FONT_NAME, size=size, bold=bold, color=color)

    @staticmethod
    def _fill(hex_color: str) -> PatternFill:
        """Return a solid fill of the given hex colour."""
        return PatternFill("solid", fgColor=hex_color)

    @staticmethod
    def _align(
        horizontal: str = "left", wrap: bool = False
    ) -> Alignment:
        """Return an alignment (vertically centred) with optional wrapping."""
        return Alignment(
            horizontal=horizontal, vertical="center", wrap_text=wrap
        )

    def _title_banner(self, ws: Worksheet, row: int, span: int, text: str) -> None:
        """Write a merged charcoal title banner across ``span`` columns."""
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=span
        )
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = self._font(14, True, Palette.WHITE)
        cell.fill = self._fill(Palette.CHARCOAL)
        cell.alignment = self._align("center")
        ws.row_dimensions[row].height = 26

    def _section_bar(self, ws: Worksheet, row: int, span: int, text: str) -> None:
        """Write a merged steel-slate section bar across ``span`` columns."""
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=span
        )
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = self._font(12, True, Palette.WHITE)
        cell.fill = self._fill(Palette.STEEL)
        cell.alignment = self._align("left")
        ws.row_dimensions[row].height = 20

    def _metadata_row(self, ws: Worksheet, row: int) -> None:
        """Write the DATE / SHIFT / APPROVED BY metadata strip."""
        pairs = [
            (1, "DATE:", self.report.get("date", "")),
            (4, "SHIFT:", "Day Shift"),
            (7, "APPROVED BY:", self.report.get("prepared_by") or "Project Controls"),
        ]
        for col, label, value in pairs:
            label_cell = ws.cell(row=row, column=col, value=label)
            label_cell.font = self._font(10, True, Palette.STEEL)
            label_cell.alignment = self._align("right")
            value_cell = ws.cell(row=row, column=col + 1, value=value)
            value_cell.font = self._font(10, False)
            value_cell.alignment = self._align("left")

    def _header_row(self, ws: Worksheet, row: int, headers: list[str]) -> None:
        """Write a muted-slate table-header row with borders."""
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=text)
            cell.font = self._font(10, True, Palette.WHITE)
            cell.fill = self._fill(Palette.MUTED)
            cell.alignment = self._align("center", wrap=True)
            cell.border = self._border
        ws.row_dimensions[row].height = 28

    # column kind: "text" (left), "center", "num", "num1" (1 decimal)
    def _data_row(
        self,
        ws: Worksheet,
        row: int,
        values: list,
        kinds: list[str],
        zebra: bool = False,
        wrap_cols: Optional[set[int]] = None,
    ) -> None:
        """Write one data row applying alignment, number formats and zebra fill."""
        wrap_cols = wrap_cols or set()
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = self._font(10, False)
            cell.border = self._border
            kind = kinds[col - 1] if col - 1 < len(kinds) else "text"
            if kind == "num":
                cell.alignment = self._align("right")
                cell.number_format = "#,##0"
            elif kind == "num1":
                cell.alignment = self._align("right")
                cell.number_format = "#,##0.0"
            elif kind == "center":
                cell.alignment = self._align("center")
            else:
                cell.alignment = self._align("left", wrap=col in wrap_cols)
            if zebra:
                cell.fill = self._fill(Palette.ZEBRA)

    def _summary_row(
        self,
        ws: Worksheet,
        row: int,
        values: list,
        kinds: list[str],
    ) -> None:
        """Write a bold summary/total row with light fill and double bottom."""
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = self._font(10, True)
            cell.fill = self._fill(Palette.SUMMARY)
            cell.border = Border(
                left=self._thin,
                right=self._thin,
                top=self._thin,
                bottom=self._double,
            )
            kind = kinds[col - 1] if col - 1 < len(kinds) else "text"
            if kind == "num":
                cell.alignment = self._align("right")
                cell.number_format = "#,##0"
            elif kind == "num1":
                cell.alignment = self._align("right")
                cell.number_format = "#,##0.0"
            else:
                cell.alignment = self._align("left")

    def _status_fill(self, cell, value: str) -> None:
        """Apply a soft status highlight (fill + text colour) to a cell."""
        mapping = {
            "working": (Palette.MINT, Palette.MINT_TEXT),
            "closed": (Palette.MINT, Palette.MINT_TEXT),
            "yes": (Palette.MINT, Palette.MINT_TEXT),
            "breakdown": (Palette.CRIMSON, Palette.CRIMSON_TEXT),
            "open": (Palette.CRIMSON, Palette.CRIMSON_TEXT),
            "high": (Palette.CRIMSON, Palette.CRIMSON_TEXT),
            "maintenance": (Palette.AMBER, Palette.AMBER_TEXT),
            "in progress": (Palette.AMBER, Palette.AMBER_TEXT),
            "medium": (Palette.AMBER, Palette.AMBER_TEXT),
            "idle": (Palette.SLATE_LIGHT, Palette.SLATE_LIGHT_TEXT),
            "standby": (Palette.SLATE_LIGHT, Palette.SLATE_LIGHT_TEXT),
            "low": (Palette.BLUE, Palette.BLUE_TEXT),
        }
        key = str(value).strip().lower()
        if key in mapping:
            fill_hex, text_hex = mapping[key]
            cell.fill = self._fill(fill_hex)
            cell.font = self._font(10, True, text_hex)

    def _discipline_tint(self, cell, discipline: str) -> None:
        """Tint a discipline label cell (civil/mechanical/electrical)."""
        tints = {
            "civil": (Palette.BLUE, Palette.BLUE_TEXT),
            "mechanical": (Palette.AMBER, Palette.AMBER_TEXT),
            "electrical": (Palette.TEAL, Palette.TEAL_TEXT),
        }
        fill_hex, text_hex = tints.get(
            discipline.strip().lower(), (Palette.SLATE_LIGHT, Palette.SLATE_LIGHT_TEXT)
        )
        cell.fill = self._fill(fill_hex)
        cell.font = self._font(10, True, text_hex)

    def _autofit(
        self, ws: Worksheet, desc_cols: Optional[set[int]] = None, cap: int = 38
    ) -> None:
        """Auto-size columns from non-merged content with +3 padding / caps."""
        desc_cols = desc_cols or set()
        covered: set[tuple[int, int]] = set()
        for rng in ws.merged_cells.ranges:
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    covered.add((r, c))

        widths: dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or (cell.row, cell.column) in covered:
                    continue
                length = max(len(s) for s in str(cell.value).split("\n"))
                widths[cell.column] = max(widths.get(cell.column, 0), length)

        for col, width in widths.items():
            letter = get_column_letter(col)
            if col in desc_cols:
                ws.column_dimensions[letter].width = min(width + 3, cap)
            else:
                ws.column_dimensions[letter].width = width + 3

    # --- sheets ---------------------------------------------------------
    def build_labor_sheet(self) -> None:
        """SHEET 1: Labor & Progress Log."""
        ws = self.workbook.active
        ws.title = "Labor & Progress Log"
        ws.sheet_view.showGridLines = True
        span = 8

        self._title_banner(
            ws, 1, span, "DAILY SITE PROGRESS & HEADCOUNT MATRIX"
        )
        self._metadata_row(ws, 2)
        self._section_bar(
            ws, 4, span, "1. DISCIPLINE HEADCOUNT & FORCE STRENGTH"
        )
        self._header_row(
            ws,
            5,
            [
                "Discipline",
                "Subcontractor / Force ID",
                "Supervisors / Eng.",
                "Foremen",
                "Skilled Labor",
                "Helpers",
                "Total Headcount",
                "Assigned Zone",
            ],
        )

        headcount = [
            ("Civil", "Civil Sub A / CIV-FORCE-01", 2, 4, 18, 12, "Zone A - Foundations"),
            ("Civil", "Civil Sub B / CIV-FORCE-02", 1, 3, 14, 9, "Zone B - Substructure"),
            ("Civil", "Civil Sub C / CIV-FORCE-03", 1, 2, 10, 6, "Zone C - Drainage"),
            ("Mechanical", "Mech Sub A / MEC-FORCE-01", 2, 3, 12, 8, "Pump House"),
            ("Mechanical", "Mech Sub B / MEC-FORCE-02", 1, 2, 9, 5, "Pipe Rack 04"),
            ("Mechanical", "Mech Sub C / MEC-FORCE-03", 1, 2, 7, 4, "Tank Farm"),
            ("Electrical", "Elec Sub A / ELE-FORCE-01", 2, 3, 10, 6, "Substation 01"),
            ("Electrical", "Elec Sub B / ELE-FORCE-02", 1, 2, 8, 5, "Cable Trench N"),
            ("Electrical", "Elec Sub C / ELE-FORCE-03", 1, 1, 6, 3, "Control Room"),
        ]
        kinds = ["text", "text", "num", "num", "num", "num", "num", "center"]
        first_data = 6
        for i, (disc, force, sup, fore, skill, help_, zone) in enumerate(headcount):
            row = first_data + i
            total = f"=SUM(C{row}:F{row})"
            self._data_row(
                ws,
                row,
                [disc, force, sup, fore, skill, help_, total, zone],
                kinds,
                zebra=(i % 2 == 1),
            )
            self._discipline_tint(ws.cell(row=row, column=1), disc)
        last_data = first_data + len(headcount) - 1  # 14

        total_row = last_data + 1  # 15
        totals = ["TOTAL SITE FORCE STRENGTH", ""]
        for col in range(3, 8):  # C..G
            letter = get_column_letter(col)
            totals.append(f"=SUM({letter}{first_data}:{letter}{last_data})")
        totals.append("")
        self._summary_row(
            ws,
            total_row,
            totals,
            ["text", "text", "num", "num", "num", "num", "num", "text"],
        )
        ws.merge_cells(
            start_row=total_row, start_column=1, end_row=total_row, end_column=2
        )

        prog_bar = total_row + 2  # 17
        self._section_bar(
            ws, prog_bar, span, "2. PROGRESS EXECUTION & MATERIAL SARFIYAT LOG"
        )
        prog_header = prog_bar + 1  # 18
        self._header_row(
            ws,
            prog_header,
            [
                "Discipline",
                "Work Location / Area",
                "Description of Daily Activities / Milestones Executed",
                "Key Materials Consumed Today",
                "Site Constraints & Delays",
                "Target Plan for Tomorrow",
            ],
        )
        progress = [
            (
                "Civil",
                "Zone A - Foundations",
                "Completed pouring of 120 m3 C30 concrete for the raft "
                "foundation at grid 1-4; curing initiated.",
                "Ready-mix concrete (120 m3), rebar (8 t)",
                "Concrete truck arrived 45 min late.",
                "Strip formwork and start column starter bars.",
            ),
            (
                "Civil",
                "Zone C - Drainage",
                "Excavated and laid 60 m of DN400 stormwater pipe; bedding "
                "compacted.",
                "DN400 pipe (60 m), granular bedding (40 m3)",
                "Minor groundwater ingress managed by pumping.",
                "Backfill trench and reinstate access road.",
            ),
            (
                "Mechanical",
                "Pipe Rack 04",
                "Erected 18 m of pipe rack steel and aligned first spool "
                "run.",
                "Structural steel (6 t), bolts, gaskets",
                "Awaiting QC release for welding.",
                "Continue spool installation on Level 2.",
            ),
            (
                "Mechanical",
                "Tank Farm",
                "Hydrotested Tank T-102; no leaks recorded at test "
                "pressure.",
                "Test water (35 m3), blind flanges",
                "None.",
                "Drain tank and prepare for internal coating.",
            ),
            (
                "Electrical",
                "Substation 01",
                "Pulled 320 m of MV cable and dressed cable trays in the "
                "switchroom.",
                "MV cable (320 m), cable cleats, glands",
                "Trench congestion delayed pulling by 1 hr.",
                "Terminate cables and begin continuity testing.",
            ),
        ]
        prog_kinds = ["text", "text", "text", "text", "text", "text"]
        for i, vals in enumerate(progress):
            row = prog_header + 1 + i
            self._data_row(
                ws,
                row,
                list(vals),
                prog_kinds,
                zebra=(i % 2 == 1),
                wrap_cols={3, 4, 5, 6},
            )
            self._discipline_tint(ws.cell(row=row, column=1), vals[0])
            ws.row_dimensions[row].height = 46

        self._autofit(ws, desc_cols={3, 4, 5, 6})

    def build_machinery_sheet(self) -> None:
        """SHEET 2: Machinery & Equipment Log."""
        ws = self.workbook.create_sheet("Machinery & Equipment Log")
        ws.sheet_view.showGridLines = True
        span = 12

        self._title_banner(ws, 1, span, "DAILY MACHINERY & FLEET OPERATIONAL LOG")
        self._metadata_row(ws, 2)
        self._header_row(
            ws,
            4,
            [
                "Equipment Code",
                "Equipment Description",
                "Operator Name",
                "Start Meter (hrs/km)",
                "End Meter (hrs/km)",
                "Gross Hours",
                "Effective Work (hrs)",
                "Idle Time (hrs)",
                "Down Time (hrs)",
                "Daily Fuel Delivery (L)",
                "Current Status",
                "Activity Remarks",
            ],
        )

        fleet = [
            ("EXC-01", "Crawler Excavator 30T", "A. Yilmaz", 1450.5, 1459.0, 7.5, 1.0, 0.0, 180, "Working", "Trenching at Zone B"),
            ("DOZ-01", "Bulldozer D8", "M. Demir", 980.0, 986.5, 5.5, 1.0, 0.0, 150, "Working", "Grading haul road"),
            ("KAM-01", "Dump Truck 20 m3", "H. Kaya", 760.0, 767.5, 6.0, 1.5, 0.0, 120, "Working", "Spoil removal to stockpile"),
            ("LDR-01", "Wheel Loader", "S. Ak", 760.0, 762.0, 1.5, 0.5, 0.0, 60, "Idle", "Awaiting trucks"),
            ("CRN-01", "Mobile Crane 80T", "R. Sahin", 540.0, 544.0, 3.0, 1.0, 0.0, 90, "Working", "Lifting precast units"),
            ("GEN-01", "Generator 250 kVA", "N/A", 3200.0, 3212.0, 12.0, 0.0, 0.0, 300, "Working", "Site temporary power"),
            ("CMP-01", "Air Compressor", "N/A", 410.0, 410.0, 0.0, 0.0, 6.0, 0, "Breakdown", "Hose rupture - awaiting repair"),
            ("GRD-01", "Motor Grader", "T. Oz", 615.0, 619.5, 4.0, 0.5, 0.0, 80, "Maintenance", "250 hr scheduled service"),
            ("TLH-01", "Telehandler 4T", "B. Can", 305.0, 309.0, 3.5, 0.5, 0.0, 45, "Working", "Material handling at racks"),
        ]
        kinds = [
            "text", "text", "text", "num1", "num1", "num1",
            "num1", "num1", "num1", "num", "center", "text",
        ]
        first = 5
        for i, item in enumerate(fleet):
            row = first + i
            code, desc, op, start, end, eff, idle, down, fuel, status, remark = item
            gross = f"=E{row}-D{row}"
            self._data_row(
                ws,
                row,
                [code, desc, op, start, end, gross, eff, idle, down, fuel, status, remark],
                kinds,
                zebra=(i % 2 == 1),
                wrap_cols={12},
            )
            self._status_fill(ws.cell(row=row, column=11), status)
        last = first + len(fleet) - 1  # 13

        total_row = last + 1  # 14
        summary = ["TOTAL FLEET SUMMARY", "", "", "", ""]
        for col in (6, 7, 8, 9, 10):  # Gross, Effective, Idle, Down, Fuel
            letter = get_column_letter(col)
            summary.append(f"=SUM({letter}{first}:{letter}{last})")
        summary.extend(["", ""])
        self._summary_row(
            ws,
            total_row,
            summary,
            [
                "text", "text", "text", "text", "text", "num1",
                "num1", "num1", "num1", "num", "text", "text",
            ],
        )
        ws.merge_cells(
            start_row=total_row, start_column=1, end_row=total_row, end_column=5
        )
        self._autofit(ws, desc_cols={12})

    def build_hse_sheet(self) -> None:
        """SHEET 3: HSE & Security Register."""
        ws = self.workbook.create_sheet("HSE & Security Register")
        ws.sheet_view.showGridLines = True
        span = 8

        self._title_banner(
            ws,
            1,
            span,
            "DAILY SITE HEALTH, SAFETY, ENVIRONMENTAL & ACCESS CONTROL REGISTER",
        )
        self._metadata_row(ws, 2)
        self._section_bar(ws, 4, span, "1. HSE SITE OBSERVATIONS & OPEN ISSUES")
        self._header_row(
            ws,
            5,
            [
                "Obs. ID",
                "Location / Zone",
                "Observation Details / Hazard Identified",
                "Category Type",
                "Risk Severity Level",
                "Corrective Action Required",
                "Responsible Party",
                "Current Status",
            ],
        )
        observations = [
            ("OBS-01", "Zone B Scaffold", "Scaffolding platform missing mid-rail and toe-board at 4 m height.", "Work at Height", "High", "Install guardrails and toe-boards; tag out until rectified.", "Civil Sub A", "Open"),
            ("OBS-02", "Tank Farm", "Two gas cylinders stored unchained near a walkway.", "Hazardous Materials", "Medium", "Secure cylinders upright in a chained rack.", "Mech Sub B", "In Progress"),
            ("OBS-03", "Substation 01", "Temporary power cable lying across the access road.", "Electrical", "Medium", "Route cable overhead or in a covered trench.", "Elec Sub A", "In Progress"),
            ("OBS-04", "Zone A Rebar Yard", "Housekeeping: offcuts and debris near the rebar yard.", "Housekeeping", "Low", "Clear debris; assign a daily housekeeping crew.", "Civil Sub B", "Closed"),
            ("OBS-05", "Main Gate", "Worker observed without safety glasses in an active zone.", "PPE", "Low", "Toolbox talk on PPE; issue safety glasses.", "HSE Officer", "Closed"),
        ]
        obs_kinds = ["center", "text", "text", "center", "center", "text", "text", "center"]
        for i, vals in enumerate(observations):
            row = 6 + i
            self._data_row(
                ws, row, list(vals), obs_kinds, zebra=(i % 2 == 1), wrap_cols={3, 6}
            )
            self._status_fill(ws.cell(row=row, column=5), vals[4])
            self._status_fill(ws.cell(row=row, column=8), vals[7])
            ws.row_dimensions[row].height = 32

        access_bar = 12
        self._section_bar(
            ws, access_bar, span, "2. SITE ACCESS, LOGISTICS & VISITOR MANAGEMENT"
        )
        self._header_row(
            ws,
            access_bar + 1,
            [
                "Pass ID",
                "Visitor Name / Carrier Company",
                "Purpose of Visit",
                "Site Host Name",
                "Time In",
                "Time Out",
                "Access Classification",
                "Induction Verified?",
            ],
        )
        access = [
            ("VP-1001", "Acme Client Audit Team", "Quality audit", "PM - J. Rivera", "09:00", "12:30", "VIP Access", "Yes"),
            ("VP-1002", "BetonPlus Ready-Mix", "Concrete delivery (120 m3)", "Civil Foreman", "06:15", "08:00", "Logistics Delivery", "Yes"),
            ("VP-1003", "City Authority Inspector", "Statutory inspection", "PM - J. Rivera", "13:30", "15:00", "Authority", "Yes"),
            ("VP-1004", "SteelCo Carrier", "Rebar delivery (8 t)", "Stores In-charge", "10:00", "11:15", "Logistics Delivery", "Yes"),
        ]
        access_kinds = ["center", "text", "text", "text", "center", "center", "center", "center"]
        access_tints = {
            "vip access": (Palette.AMBER, Palette.AMBER_TEXT),
            "logistics delivery": (Palette.SLATE_LIGHT, Palette.SLATE_LIGHT_TEXT),
            "authority": (Palette.BLUE, Palette.BLUE_TEXT),
        }
        for i, vals in enumerate(access):
            row = access_bar + 2 + i
            self._data_row(
                ws, row, list(vals), access_kinds, zebra=(i % 2 == 1)
            )
            cls_cell = ws.cell(row=row, column=7)
            tint = access_tints.get(str(vals[6]).strip().lower())
            if tint:
                cls_cell.fill = self._fill(tint[0])
                cls_cell.font = self._font(10, True, tint[1])
            self._status_fill(ws.cell(row=row, column=8), vals[7])

        self._autofit(ws, desc_cols={3, 6})

    def build(self) -> Workbook:
        """Build all three sheets and return the workbook."""
        self.build_labor_sheet()
        self.build_machinery_sheet()
        self.build_hse_sheet()
        return self.workbook

    def save(self, output_path: str) -> str:
        """Build and save the workbook to ``output_path``; return the path."""
        self.build()
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        self.workbook.save(output_path)
        return output_path


def build_daily_report_excel(output_path: str, report: Optional[dict] = None) -> str:
    """Render the styled daily-report workbook for ``report`` to ``output_path``."""
    return DailyReportExcelBuilder(report).save(output_path)


if __name__ == "__main__":
    sample = {"date": "2026-06-08", "prepared_by": "J. Rivera", "project_name": "Metro Station A"}
    path = build_daily_report_excel("daily_report_sample.xlsx", sample)
    print(f"Saved sample workbook: {path}")
